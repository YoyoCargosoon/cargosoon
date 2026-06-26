import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")


DOWNLOADS_DIR = Path(r"C:\Users\YOYO\Downloads")
OUTPUT_PATH = Path("src/data/amazonWarehouseCatalog.generated.json")


SEED_RECORDS = [
    {"code": "ONT8", "market": "US", "country": "United States", "address": "24300 Nandina Ave., Moreno Valley, CA 92551", "source": "RN 美国专线价格表.xlsx / FBA仓库地址表"},
    {"code": "ONT6", "market": "US", "country": "United States", "address": "24208 San Michele Rd., Moreno Valley, CA 92551", "source": "RN 美国专线价格表.xlsx / FBA仓库地址表"},
    {"code": "ONT5", "market": "US", "country": "United States", "address": "2020 E. Central Ave., Southgate Building 4, San Bernardino, CA 92408", "source": "RN 美国专线价格表.xlsx / FBA仓库地址表"},
    {"code": "ONT4", "market": "US", "country": "United States", "address": "1910 E Central Ave., San Bernardino, CA 92408", "source": "RN 美国专线价格表.xlsx / FBA仓库地址表"},
    {"code": "ONT2", "market": "US", "country": "United States", "address": "1910 E Central Ave., San Bernardino, CA 92408", "source": "RN 美国专线价格表.xlsx / FBA仓库地址表"},
    {"code": "SBD2", "market": "US", "country": "United States", "address": "1494 S Waterman Ave, San Bernardino, CA 92408", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "PHX3", "market": "US", "country": "United States", "address": "6835 West Buckeye Road, Phoenix, AZ 85043", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "PHX5", "market": "US", "country": "United States", "address": "16980 W Commerce Drive, Goodyear, AZ 85338", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "PHX6", "market": "US", "country": "United States", "address": "4750 West Mohave St, Phoenix, AZ 85043", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "PHX7", "market": "US", "country": "United States", "address": "800 N 75th Ave, Phoenix, AZ 85043", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "PHX8", "market": "US", "country": "United States", "address": "800 N 75th Ave, Phoenix, AZ 85043", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "TFC1", "market": "US", "country": "United States", "address": "5050 West Mohave St, Phoenix, AZ 85043", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "MOB5", "market": "US", "country": "United States", "address": "6735 Trippel Road, Mobile, AL 36619", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "DPX1", "market": "US", "country": "United States", "address": "500 S. 48th Street, Phoenix, AZ 85034", "source": "HYJM 美加专线价格表.xlsx / 美国FBA仓"},
    {"code": "PSC2", "market": "US", "country": "United States", "address": "1351 S Road 40 E, Pasco, WA 99301", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "GEG2", "market": "US", "country": "United States", "address": "18007 E Garland Ave, Spokane Valley, WA 99216-6102", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "BFI1", "market": "US", "country": "United States", "address": "1800 140th Avenue E., Sumner, WA 98390", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "BFI7", "market": "US", "country": "United States", "address": "1901 140th Ave E, Sumner, WA 98390", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "BFI3", "market": "US", "country": "United States", "address": "2700 Center Drive, DuPont, WA 98327-9607", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "BFI9", "market": "US", "country": "United States", "address": "3230 International Pl, DuPont, WA 98327-7707", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "BFI4", "market": "US", "country": "United States", "address": "21005 64th Ave S, Kent, WA 98032", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "PDX7", "market": "US", "country": "United States", "address": "4775 Depot Ct SE, Salem, OR 97317", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "PDX6", "market": "US", "country": "United States", "address": "15000 N. Lombard St, Portland, OR 97203", "source": "HMLX 美国 加拿大 海派 海卡 FBA .xlsx / FBA仓库代码表"},
    {"code": "YVR2", "market": "CA", "country": "Canada", "address": "450 Derwent St., Delta, BC V3M5Y9", "source": "HYJM 美加专线价格表.xlsx / 加拿大FBA海外仓分区"},
    {"code": "YVR3", "market": "CA", "country": "Canada", "address": "109 Braid St, New Westminster, BC V3L5H4", "source": "HYJM 美加专线价格表.xlsx / 加拿大FBA海外仓分区"},
    {"code": "YVR4", "market": "CA", "country": "Canada", "address": "4189 Salish Sea Way, Tsawwassen, BC V4M0B9", "source": "HYJM 美加专线价格表.xlsx / 加拿大FBA海外仓分区"},
    {"code": "YXX1", "market": "CA", "country": "Canada", "address": "19300 Airport Way, Pitt Meadows, BC V3Y2M5", "source": "HYJM 美加专线价格表.xlsx / 加拿大FBA海外仓分区"},
]


FALLBACK_RECORDS = [
    ("LAX9", "US", "United States", "Fontana, California, United States"),
    ("SBD1", "US", "United States", "Bloomington, California, United States"),
    ("JFK8", "US", "United States", "Staten Island, New York, United States"),
    ("ORD2", "US", "United States", "Channahon, Illinois, United States"),
    ("FWA4", "US", "United States", "Fort Wayne, Indiana, United States"),
    ("YYZ7", "CA", "Canada", "Caledon / Bolton, Ontario, Canada"),
    ("YOW3", "CA", "Canada", "Ottawa, Ontario, Canada"),
    ("BHX4", "UK", "United Kingdom", "Coventry, England, United Kingdom"),
    ("MAN1", "UK", "United Kingdom", "Manchester / Bolton, England, United Kingdom"),
    ("EMA1", "UK", "United Kingdom", "Derby / Coalville area, England, United Kingdom"),
    ("DTM2", "DE", "Germany", "Dortmund, North Rhine-Westphalia, Germany"),
    ("FRA7", "DE", "Germany", "Frankenthal, Rhineland-Palatinate, Germany"),
    ("LEJ1", "DE", "Germany", "Leipzig, Saxony, Germany"),
    ("BVA1", "FR", "France", "France Amazon Fulfillment Center"),
    ("MRS1", "FR", "France", "France Amazon Fulfillment Center"),
    ("MXP5", "IT", "Italy", "Italy Amazon Fulfillment Center"),
    ("TRN1", "IT", "Italy", "Italy Amazon Fulfillment Center"),
    ("BCN1", "ES", "Spain", "Spain Amazon Fulfillment Center"),
    ("MAD4", "ES", "Spain", "Spain Amazon Fulfillment Center"),
    ("RTM1", "NL", "Netherlands", "Netherlands Amazon Fulfillment Center"),
    ("AMS7", "NL", "Netherlands", "Netherlands Amazon Fulfillment Center"),
    ("NRT5", "JP", "Japan", "Narita / Greater Tokyo Area, Japan"),
    ("KIX6", "JP", "Japan", "Osaka / Kansai Area, Japan"),
]


def is_valid_code(value):
    text = str(value or "").strip().upper()
    return bool(re.fullmatch(r"[A-Z0-9-]{3,10}", text)) and any(ch.isdigit() for ch in text)


def find_file(prefix):
    return next((path for path in DOWNLOADS_DIR.glob("*.xlsx") if path.name.startswith(prefix)), None)


def add_record(records, code, market, country, address, source):
    code = str(code or "").strip().upper()
    if not is_valid_code(code):
      return
    address = str(address or "").strip()
    entry = {
        "code": code,
        "market": market,
        "country": country,
        "address": address,
        "source": source,
    }
    if code not in records or (address and len(address) > len(records[code].get("address", ""))):
        records[code] = entry


def import_hmlx(records):
    path = find_file("HMLX")
    if not path:
        return
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[19]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        country, code, _, trailing, address = row[:5]
        if not code or not address:
            continue
        if country == "美国":
            market, country_name = "US", "United States"
        elif country == "加拿大":
            market, country_name = "CA", "Canada"
        else:
            continue
        full_address = str(address).strip()
        trailing_text = str(trailing or "").strip()
        if trailing_text and trailing_text not in full_address:
            full_address = f"{full_address}, {trailing_text}"
        add_record(records, code, market, country_name, full_address, f"{path.name} / {sheet.title}")


def import_rn(records):
    path = find_file("RN ")
    if not path:
        return
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[19]
    for row in sheet.iter_rows(min_row=3, values_only=True):
        code, state, city, address1, address2, zip_code = row[:6]
        if not code:
            continue
        full_address = ", ".join(
            str(item).strip() for item in [address1, address2, city, state, zip_code] if item not in (None, "")
        )
        add_record(records, code, "US", "United States", full_address, f"{path.name} / {sheet.title}")


def import_hyjm(records):
    path = find_file("HYJM")
    if not path:
        return
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    us_sheet = workbook.worksheets[31]
    for row in us_sheet.iter_rows(min_row=4, values_only=True):
        code, address, city, state, zip_code = row[:5]
        if not code:
            continue
        full_address = ", ".join(
            str(item).strip() for item in [address, city, state, zip_code] if item not in (None, "")
        )
        add_record(records, code, "US", "United States", full_address, f"{path.name} / {us_sheet.title}")

    ca_sheet = workbook.worksheets[44]
    for row in ca_sheet.iter_rows(min_row=4, values_only=True):
        values = [str(item).strip() if item is not None else "" for item in row[:6]]
        code = ""
        if is_valid_code(values[0]):
            code = values[0]
            full_address = ", ".join(value for value in values[1:5] if value)
        elif is_valid_code(values[1]):
            code = values[1]
            full_address = ", ".join(value for value in values[2:6] if value)
        else:
            continue
        add_record(records, code, "CA", "Canada", full_address, f"{path.name} / {ca_sheet.title}")


def main():
    records = {item["code"]: item for item in SEED_RECORDS}

    import_hmlx(records)
    import_rn(records)
    import_hyjm(records)

    for code, market, country, address in FALLBACK_RECORDS:
        add_record(records, code, market, country, address, "manual fallback / coarse location")

    catalog = {
        "updated_at": "2026-06-25",
        "description": "Amazon warehouse catalog for quote page autocomplete and address autofill.",
        "records": sorted(records.values(), key=lambda item: (item["market"], item["code"])),
    }

    OUTPUT_PATH.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Records: {len(catalog['records'])}")


if __name__ == "__main__":
    main()
