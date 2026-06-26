import json
import re
from pathlib import Path

import openpyxl


OUTPUT_PATH = Path("src/data/ddpRateCatalog.generated.json")
DOWNLOADS = Path(r"C:\Users\YOYO\Downloads")


def find_file(prefix):
    return next((p for p in DOWNLOADS.glob("*.xlsx") if p.name.startswith(prefix)), None)


def safe_text(value):
    return str(value).strip() if value is not None else ""


def as_cells(row, width):
    return [safe_text(v) for v in row[:width]]


def to_float(value):
    cleaned = safe_text(value).replace("+派送费", "").replace(",", "")
    try:
        return float(cleaned)
    except Exception:
        return None


def add_rate(records, **kwargs):
    if kwargs.get("price") is None:
        return
    kwargs["rate_id"] = f"ddp-{len(records) + 1}"
    kwargs["service_type"] = "ddp"
    records.append(kwargs)


def split_lines(value):
    return [item.strip() for item in safe_text(value).replace("、", "/").replace("\n", "/").split("/") if item.strip()]


def extract_codes(text):
    return re.findall(r"\b[A-Z]{2,}\d(?:-\d+)?\b|\bX[A-Z]{2,}\d\b|\b4PX-\d+\b|\bPRTO\b|\bXCAB\b", safe_text(text).upper())


def normalize_zone_label(text):
    cleaned = safe_text(text).replace("\n", " ").replace("  ", " ").strip()
    return cleaned


def add_weight_bands(records, *, service_mode, market, zone, channel_name, source_file, source_sheet, transit, weights, prices, extra=None):
    extra = extra or {}
    for (label, min_weight, max_weight), price in zip(weights, prices):
        add_rate(
            records,
            service_mode=service_mode,
            market=market,
            zone=zone,
            band_label=label,
            min_weight_kg=min_weight,
            max_weight_kg=max_weight,
            currency="RMB",
            price=to_float(price),
            channel_name=channel_name,
            transit=transit,
            source_file=source_file,
            source_sheet=source_sheet,
            **extra,
        )


def parse_rn(records):
    path = find_file("RN ")
    if not path:
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sea_sheet = wb.worksheets[2]
    current_channel = ""
    zone_map = {"美东": "US East (ZIP 0-3)", "美中": "US Central (ZIP 4-7)", "美西": "US West (ZIP 8-9)"}
    for row in sea_sheet.iter_rows(min_row=1, max_row=120, values_only=True):
        cells = as_cells(row, 10)
        if cells[0] and "船期" in cells[1]:
            current_channel = cells[0]
        label = cells[0].replace("\n", "")
        for prefix, zone in zone_map.items():
            if label.startswith(prefix):
                add_weight_bands(
                    records,
                    service_mode="sea_ddp",
                    market="United States",
                    zone=zone,
                    channel_name=current_channel or "RN US Sea DDP",
                    source_file=path.name,
                    source_sheet=sea_sheet.title,
                    transit=cells[5],
                    weights=[("12-70KG", 12, 70), ("71-99KG", 71, 99), ("100KG+", 100, None)],
                    prices=[cells[2], cells[3], cells[4]],
                )

    air_sheet = wb.worksheets[14]
    current_channel = ""
    for row in air_sheet.iter_rows(min_row=1, max_row=80, values_only=True):
        cells = as_cells(row, 8)
        if "统配" in cells[0] or cells[0].startswith("空"):
            current_channel = cells[0]
        zone = ""
        if cells[0].startswith("美西"):
            zone = "US West (ZIP 8-9)"
        elif cells[0].startswith("美中"):
            zone = "US Central (ZIP 4-7)"
        elif cells[0].startswith("美东"):
            zone = "US East (ZIP 0-3)"
        if not zone:
            continue
        add_weight_bands(
            records,
            service_mode="air_ddp",
            market="United States",
            zone=zone,
            channel_name=current_channel or "RN US Air DDP",
            source_file=path.name,
            source_sheet=air_sheet.title,
            transit=cells[5],
            weights=[("12-20KG", 12, 20), ("21-100KG", 21, 100), ("101KG+", 101, None)],
            prices=[cells[2], cells[3], cells[4]],
        )


def parse_ythq(records):
    path = find_file("YTHQ")
    if not path:
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sea_sheet = wb.worksheets[2]
    current_channel = ""
    transit = ""
    zone_map = {"美西": "US West (ZIP 8-9)", "美中": "US Central (ZIP 4-7)", "美东": "US East (ZIP 0-3)"}
    for row in sea_sheet.iter_rows(min_row=1, max_row=140, values_only=True):
        cells = as_cells(row, 10)
        if ("海派" in cells[0] or "海运" in cells[0]) and any(token in cells[1] for token in zone_map):
            current_channel = cells[0]
            transit = cells[7]
            zone = next((label for token, label in zone_map.items() if token in cells[1]), "")
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="United States",
                zone=zone,
                channel_name=current_channel,
                source_file=path.name,
                source_sheet=sea_sheet.title,
                transit=transit,
                weights=[("12-99KG", 12, 99), ("100KG+", 100, None)],
                prices=[cells[2], cells[3]],
            )
        elif any(cells[0].startswith(prefix) for prefix in zone_map):
            zone = next((label for token, label in zone_map.items() if cells[0].startswith(token)), "")
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="United States",
                zone=zone,
                channel_name=current_channel or "YTHQ US Sea DDP",
                source_file=path.name,
                source_sheet=sea_sheet.title,
                transit=transit,
                weights=[("12-99KG", 12, 99), ("100KG+", 100, None)],
                prices=[cells[1], cells[2]],
            )

    uk_express = wb.worksheets[19]
    mode = ""
    for row in uk_express.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 10)
        if "PVA" in cells[0]:
            mode = "uk_pva"
        elif "包税" in cells[0]:
            mode = "uk_taxed"
        elif cells[0] == "英国":
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[1]),
                channel_name="UK Sea Express PVA" if mode == "uk_pva" else "UK Sea DDP",
                source_file=path.name,
                source_sheet=uk_express.title,
                transit=cells[5],
                weights=[("21-99KG", 21, 99), ("100KG+", 100, None)],
                prices=[cells[2], cells[3]],
                extra={"delivery_type": "express"},
            )

    uk_truck = wb.worksheets[20]
    band_header = [("100-499KG", 100, 499), ("500-999KG", 500, 999), ("1000KG+", 1000, None)]
    for row in uk_truck.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 9)
        if "特价仓" in cells[0] or "所有FBA仓" in cells[0] or "Orange" in cells[0]:
            add_weight_bands(
                records,
                service_mode="truck_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[0]),
                channel_name="UK Sea Truck PVA",
                source_file=path.name,
                source_sheet=uk_truck.title,
                transit=cells[8],
                weights=band_header,
                prices=[cells[1], cells[2], cells[3]],
            )
        elif "大型海外仓" in cells[0] or "所有FBA仓" in cells[0]:
            add_rate(
                records,
                service_mode="truck_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[0]),
                band_label="100KG+",
                min_weight_kg=100,
                max_weight_kg=None,
                currency="RMB",
                price=to_float(cells[1]),
                channel_name="UK Sea Truck Taxed",
                transit=cells[8],
                source_file=path.name,
                source_sheet=uk_truck.title,
                delivery_type="truck",
            )

    eu_express = wb.worksheets[21]
    mode = ""
    for row in eu_express.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 9)
        if "递延" in cells[0]:
            mode = "pva"
        elif "包税" in cells[0]:
            mode = "taxed"
        elif cells[0] and cells[0] != "国家":
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="Europe",
                zone=normalize_zone_label(cells[0]),
                channel_name="EU Sea Express Deferred" if mode == "pva" else "EU Sea DDP Express",
                source_file=path.name,
                source_sheet=eu_express.title,
                transit=cells[6],
                weights=[("21-50KG", 21, 50), ("51-99KG", 51, 99), ("100-999KG", 100, 999), ("1000KG+", 1000, None)],
                prices=[cells[1], cells[2], cells[3], cells[4]],
                extra={"delivery_type": "express"},
            )

    eu_truck = wb.worksheets[22]
    phase = ""
    for row in eu_truck.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 10)
        if "递延卡派" in cells[1]:
            phase = "pva"
        elif "包税卡派" in cells[1]:
            phase = "taxed"
        elif cells[0] and cells[0] != "仓库代码":
            add_weight_bands(
                records,
                service_mode="truck_ddp",
                market="Europe",
                zone=normalize_zone_label(cells[0]),
                channel_name="EU Sea Truck Deferred" if phase == "pva" else "EU Sea Truck Taxed",
                source_file=path.name,
                source_sheet=eu_truck.title,
                transit=cells[9],
                weights=[("21-50KG", 21, 50), ("51-99KG", 51, 99), ("100-999KG", 100, 999), ("1000KG+", 1000, None)],
                prices=[cells[1], cells[2], cells[3], cells[4]],
                extra={"delivery_type": "truck", "warehouse_codes": extract_codes(cells[0])},
            )


def parse_hyjm(records):
    path = find_file("HYJM")
    if not path:
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    us_air_sheet = wb.worksheets[24]
    current_channel = ""
    for row in us_air_sheet.iter_rows(min_row=1, max_row=50, values_only=True):
        cells = as_cells(row, 9)
        if "体积除5000" in cells[0]:
            current_channel = "US Air B / Vol 5000"
        elif "体积除6000" in cells[0]:
            current_channel = "US Air B / Vol 6000"
        elif cells[0].startswith("邮编") or cells[0].startswith("美西自提"):
            add_weight_bands(
                records,
                service_mode="air_ddp",
                market="United States",
                zone=normalize_zone_label(cells[0]),
                channel_name=current_channel or "US Air B",
                source_file=path.name,
                source_sheet=us_air_sheet.title,
                transit=cells[6],
                weights=[("12-20KG", 12, 20), ("21-50KG", 21, 50), ("51-70KG", 51, 70), ("71-100KG", 71, 100), ("101KG+", 101, None)],
                prices=[cells[1], cells[2], cells[3], cells[4], cells[5]],
            )

    ca_air_sheet = wb.worksheets[33]
    current_channel = ""
    zone_groups = []
    for row in ca_air_sheet.iter_rows(min_row=1, max_row=30, values_only=True):
        cells = as_cells(row, 10)
        if "加拿大空派A价" in cells[1]:
            current_channel = cells[1]
        elif cells[1].startswith("体积除"):
            current_channel = f"{current_channel} / {cells[1]}"
        elif cells[1] and any(char.isdigit() for char in cells[2:6]):
            zone_label = normalize_zone_label(cells[1].split("\n")[0])
            zone_groups.append(zone_label)
            add_weight_bands(
                records,
                service_mode="air_ddp",
                market="Canada",
                zone=zone_label,
                channel_name=current_channel or "Canada Air A",
                source_file=path.name,
                source_sheet=ca_air_sheet.title,
                transit=cells[6],
                weights=[("12-20KG", 12, 20), ("21-70KG", 21, 70), ("71-100KG", 71, 100), ("101KG+", 101, None)],
                prices=[cells[2], cells[3], cells[4], cells[5]],
                extra={"warehouse_codes": extract_codes(cells[1])},
            )

    ca_sea_sheet = wb.worksheets[38]
    current_channel = ""
    for row in ca_sea_sheet.iter_rows(min_row=1, max_row=30, values_only=True):
        cells = as_cells(row, 10)
        if "加拿大海派" in cells[1]:
            current_channel = cells[1]
        elif cells[1].startswith("温哥华") or cells[1].startswith("卡尔加里") or cells[1].startswith("多伦多") or cells[1].startswith("渥太华"):
            zone_label = normalize_zone_label(cells[1].split("\n")[0])
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="Canada",
                zone=zone_label,
                channel_name=current_channel or "Canada Sea C",
                source_file=path.name,
                source_sheet=ca_sea_sheet.title,
                transit=cells[8],
                weights=[("21-50KG", 21, 50), ("51-70KG", 51, 70), ("71-100KG", 71, 100), ("101-999KG", 101, 999), ("1000KG+", 1000, None)],
                prices=[cells[2], cells[3], cells[4], cells[5], cells[6]],
                extra={"warehouse_codes": extract_codes(cells[1])},
            )

    ca_fba_sheet = wb.worksheets[39]
    current_channel = ""
    for row in ca_fba_sheet.iter_rows(min_row=1, max_row=30, values_only=True):
        cells = as_cells(row, 9)
        if "加拿大FBA海卡包税" in cells[1]:
            current_channel = cells[1]
        elif cells[1].startswith("温哥华") or cells[1].startswith("卡尔加里") or cells[1].startswith("多伦多") or cells[1].startswith("渥太华"):
            zone_label = normalize_zone_label(cells[1].split("\n")[0])
            add_weight_bands(
                records,
                service_mode="truck_ddp",
                market="Canada",
                zone=zone_label,
                channel_name=current_channel or "Canada FBA Sea Truck",
                source_file=path.name,
                source_sheet=ca_fba_sheet.title,
                transit=cells[6],
                weights=[("21-50KG", 21, 50), ("51-99KG", 51, 99), ("100KG+", 100, None)],
                prices=[cells[2], cells[3], cells[4]],
                extra={"warehouse_codes": extract_codes(cells[1]), "delivery_type": "fba_truck"},
            )


def parse_yh_pc(records):
    path = find_file("YH_PC")
    if not path:
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    eu_sheet = wb.worksheets[5]
    for row in eu_sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 10)
        if cells[0].startswith("PC-03") or ("德国" in cells[1] or "法国" in cells[1] or "波兰" in cells[1]):
            zone = normalize_zone_label(cells[1] or cells[0])
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="Europe",
                zone=zone,
                channel_name="EU Sea Courier BS",
                source_file=path.name,
                source_sheet=eu_sheet.title,
                transit=cells[9],
                weights=[("21-50KG", 21, 50), ("51-99KG", 51, 99), ("100KG+", 100, None)],
                prices=[cells[2], cells[3], cells[4]],
                extra={"delivery_type": "express"},
            )
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="Europe",
                zone=f"{zone} / self-tax",
                channel_name="EU Sea Courier Self Tax",
                source_file=path.name,
                source_sheet=eu_sheet.title,
                transit=cells[9],
                weights=[("21-50KG", 21, 50), ("51-99KG", 51, 99), ("100KG+", 100, None)],
                prices=[cells[6], cells[7], cells[8]],
                extra={"delivery_type": "express"},
            )

    uk_truck = wb.worksheets[7]
    for row in uk_truck.iter_rows(min_row=1, max_row=15, values_only=True):
        cells = as_cells(row, 10)
        if cells[0].startswith("PC-07") or cells[0].startswith("PC-08") or cells[1]:
            zone = normalize_zone_label(cells[1] or "United Kingdom")
            channel = cells[0] or "UK Sea Truck"
            add_weight_bands(
                records,
                service_mode="truck_ddp",
                market="United Kingdom",
                zone=zone,
                channel_name=channel,
                source_file=path.name,
                source_sheet=uk_truck.title,
                transit=cells[9],
                weights=[("21-50KG", 21, 50), ("51-99KG", 51, 99), ("100KG+", 100, None)],
                prices=[cells[2], cells[3], cells[4]],
                extra={"delivery_type": "truck"},
            )

    uk_sea = wb.worksheets[8]
    for row in uk_sea.iter_rows(min_row=1, max_row=12, values_only=True):
        cells = as_cells(row, 10)
        if cells[0].startswith("英国海运快递派"):
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[2]),
                channel_name=cells[0],
                source_file=path.name,
                source_sheet=uk_sea.title,
                transit=cells[6],
                weights=[("21-50KG", 21, 50), ("51-99KG", 51, 99), ("100KG+", 100, None)],
                prices=[cells[3], cells[4], cells[5]],
                extra={"delivery_type": "express"},
            )


def parse_ym(records):
    path = find_file("YM 欧洲英国海派海卡卡航铁派，超大件，纯电DG柜")
    if not path:
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    eu_sea = wb.worksheets[1]
    for row in eu_sea.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 10)
        if cells[0].startswith("J") or cells[1]:
            channel_name = cells[0] or "EU Sea"
            zone = normalize_zone_label(cells[1])
            if not zone:
                continue
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="Europe",
                zone=zone,
                channel_name=channel_name,
                source_file=path.name,
                source_sheet=eu_sea.title,
                transit="35-45 days",
                weights=[("25-70KG", 25, 70), ("71-999KG", 71, 999), ("1000KG+", 1000, None)],
                prices=[cells[2], cells[3], cells[4]],
                extra={"warehouse_codes": extract_codes(zone)},
            )

    eu_rail = wb.worksheets[3]
    for row in eu_rail.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 10)
        if cells[0].startswith("T") or cells[1]:
            channel_name = cells[0] or "EU Rail"
            zone = normalize_zone_label(cells[1])
            if not zone:
                continue
            add_weight_bands(
                records,
                service_mode="rail_ddp",
                market="Europe",
                zone=zone,
                channel_name=channel_name,
                source_file=path.name,
                source_sheet=eu_rail.title,
                transit="Rail schedule",
                weights=[("25-70KG", 25, 70), ("71-999KG", 71, 999), ("1000KG+", 1000, None)],
                prices=[cells[2], cells[3], cells[4]],
                extra={"warehouse_codes": extract_codes(zone)},
            )

    uk_sea = wb.worksheets[4]
    for row in uk_sea.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 10)
        if cells[0].startswith("E"):
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[1]),
                channel_name=cells[0],
                source_file=path.name,
                source_sheet=uk_sea.title,
                transit="UK sea schedule",
                weights=[("25-99KG", 25, 99), ("100-999KG", 100, 999), ("1000-2999KG", 1000, 2999), ("3000KG+", 3000, None)],
                prices=[cells[2], cells[3], cells[4], cells[5]],
            )
        elif cells[0].startswith("F") or "亚马逊" in cells[1]:
            add_weight_bands(
                records,
                service_mode="truck_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[1]),
                channel_name=cells[0] or "UK Sea Truck",
                source_file=path.name,
                source_sheet=uk_sea.title,
                transit="UK sea truck schedule",
                weights=[("25-99KG", 25, 99), ("100-999KG", 100, 999), ("1000-2999KG", 1000, 2999), ("3000KG+", 3000, None)],
                prices=[cells[2], cells[3], cells[4], cells[5]],
            )

    uk_rail = wb.worksheets[5]
    for row in uk_rail.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 10)
        if cells[0].startswith("T8") or cells[0].startswith("T10"):
            add_weight_bands(
                records,
                service_mode="rail_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[1]),
                channel_name=cells[0],
                source_file=path.name,
                source_sheet=uk_rail.title,
                transit="UK rail express",
                weights=[("25-70KG", 25, 70), ("71-999KG", 71, 999), ("1000KG+", 1000, None)],
                prices=[cells[2], cells[3], cells[4]],
            )
        elif cells[0].startswith("T11") or cells[0].startswith("T13"):
            add_weight_bands(
                records,
                service_mode="rail_ddp",
                market="United Kingdom",
                zone=normalize_zone_label(cells[1]),
                channel_name=cells[0],
                source_file=path.name,
                source_sheet=uk_rail.title,
                transit="UK rail truck",
                weights=[("25-70KG", 25, 70), ("71-999KG", 71, 999), ("1000KG+", 1000, None)],
                prices=[cells[2], cells[3], cells[4]],
                extra={"delivery_type": "truck"},
            )

    ym_air_path = find_file("YM 欧洲英国空派")
    if ym_air_path:
        air_wb = openpyxl.load_workbook(ym_air_path, read_only=True, data_only=True)

        eu_air = air_wb.worksheets[1]
        for row in eu_air.iter_rows(min_row=1, max_row=20, values_only=True):
            cells = as_cells(row, 10)
            if cells[0] in {"德国", "法国 比利时 卢森堡  荷兰 奥地利", "意大利  西班牙 波兰 捷克 斯洛伐克  斯洛文尼亚 丹麦 芬兰 瑞典 匈牙利 葡萄牙", "保加利亚 爱沙尼亚 拉脱维亚 立陶宛 希腊 克罗地亚"}:
                add_weight_bands(
                    records,
                    service_mode="air_ddp",
                    market="Europe",
                    zone=normalize_zone_label(cells[0]),
                    channel_name="EU Air DDP",
                    source_file=ym_air_path.name,
                    source_sheet=eu_air.title,
                    transit=cells[8],
                    weights=[("12-24KG", 12, 24), ("25-70KG", 25, 70), ("71-99KG", 71, 99), ("100-299KG", 100, 299), ("300-499KG", 300, 499), ("500KG+", 500, None)],
                    prices=[cells[2], cells[3], cells[4], cells[5], cells[6], cells[7]],
                    extra={"delivery_type": "air"},
                )

        uk_air_1 = air_wb.worksheets[2]
        uk_air_2 = air_wb.worksheets[3]
        for ws in [uk_air_1, uk_air_2]:
            for row in ws.iter_rows(min_row=1, max_row=16, values_only=True):
                cells = as_cells(row, 10)
                if cells[0] == "英国":
                    add_weight_bands(
                        records,
                        service_mode="air_ddp",
                        market="United Kingdom",
                        zone="United Kingdom",
                        channel_name=ws.title,
                        source_file=ym_air_path.name,
                        source_sheet=ws.title,
                        transit=cells[8] if len(cells) > 8 else "",
                        weights=[("12-24KG", 12, 24), ("25-70KG", 25, 70), ("71-99KG", 71, 99), ("100-299KG", 100, 299), ("300-499KG", 300, 499), ("500KG+", 500, None)],
                        prices=[cells[2], cells[3], cells[4], cells[5], cells[6], cells[7]],
                    )

        ca_air = air_wb.worksheets[6]
        current_zone = ""
        for row in ca_air.iter_rows(min_row=1, max_row=20, values_only=True):
            cells = as_cells(row, 10)
            if any(keyword in cells[0] for keyword in ["多伦多", "温哥华", "卡尔加里", "渥太华"]):
                current_zone = normalize_zone_label(cells[0])
            elif current_zone and any(to_float(item) is not None for item in cells[1:6]):
                add_weight_bands(
                    records,
                    service_mode="air_ddp",
                    market="Canada",
                    zone=current_zone,
                    channel_name="Canada Air DDP",
                    source_file=ym_air_path.name,
                    source_sheet=ca_air.title,
                    transit=cells[8] if len(cells) > 8 else "",
                    weights=[("12-24KG", 12, 24), ("25-70KG", 25, 70), ("71-99KG", 71, 99), ("100-299KG", 100, 299), ("300-499KG", 300, 499)],
                    prices=[cells[1], cells[2], cells[3], cells[4], cells[5]],
                )

        eu_air_pva = air_wb.worksheets[7]
        current_channel = ""
        for row in eu_air_pva.iter_rows(min_row=1, max_row=20, values_only=True):
            cells = as_cells(row, 10)
            if cells[0].startswith("C2") or cells[0].startswith("C4"):
                current_channel = cells[0]
            elif cells[0] in {"德国", "法国/捷克/波兰", "西班牙/意大利/葡萄牙"}:
                add_weight_bands(
                    records,
                    service_mode="air_ddp",
                    market="Europe",
                    zone=normalize_zone_label(cells[0]),
                    channel_name=current_channel or "EU Air Deferred",
                    source_file=ym_air_path.name,
                    source_sheet=eu_air_pva.title,
                    transit=cells[8],
                    weights=[("25-70KG", 25, 70), ("71-99KG", 71, 99), ("100-299KG", 100, 299), ("300-499KG", 300, 499), ("500KG+", 500, None)],
                    prices=[cells[1], cells[2], cells[3], cells[4], cells[5]],
                    extra={"delivery_type": "air"},
                )


def parse_yt(records):
    path = find_file("YT ")
    if not path:
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    uae_air = wb.worksheets[1]
    for row in uae_air.iter_rows(min_row=1, max_row=40, values_only=True):
        cells = as_cells(row, 10)
        if cells[1] in ["迪拜/沙迦", "其他区"]:
            zone = f"UAE {cells[1]}"
            add_weight_bands(
                records,
                service_mode="air_ddp",
                market="United Arab Emirates",
                zone=zone,
                channel_name=cells[0] or "UAE Air DDP",
                source_file=path.name,
                source_sheet=uae_air.title,
                transit=cells[8],
                weights=[("16-99KG", 16, 99), ("100-499KG", 100, 499), ("500-999KG", 500, 999), ("1000KG+", 1000, None)],
                prices=[cells[3], cells[4], cells[5], cells[6]],
            )

    uae_sea = wb.worksheets[2]
    for row in uae_sea.iter_rows(min_row=1, max_row=20, values_only=True):
        cells = as_cells(row, 8)
        if cells[0] in ["普货类", "牌子类", "电池类", "敏感货"]:
            add_rate(
                records,
                service_mode="sea_ddp",
                market="United Arab Emirates",
                zone=cells[0],
                band_label="0.5-5CBM",
                min_weight_kg=None,
                max_weight_kg=None,
                currency="RMB",
                price=to_float(cells[1]),
                channel_name="UAE Sea DDP",
                transit=cells[4],
                source_file=path.name,
                source_sheet=uae_sea.title,
            )
            add_rate(
                records,
                service_mode="sea_ddp",
                market="United Arab Emirates",
                zone=cells[0],
                band_label="5CBM+",
                min_weight_kg=None,
                max_weight_kg=None,
                currency="RMB",
                price=to_float(cells[2]),
                channel_name="UAE Sea DDP",
                transit=cells[4],
                source_file=path.name,
                source_sheet=uae_sea.title,
            )


def parse_yx(records):
    path = find_file("YX ")
    if not path:
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    au_air = wb.worksheets[3]
    current_channel = ""
    for row in au_air.iter_rows(min_row=1, max_row=30, values_only=True):
        cells = as_cells(row, 7)
        if "澳洲空派" in cells[0]:
            current_channel = cells[0]
        elif cells[0].endswith("区查询"):
            add_weight_bands(
                records,
                service_mode="air_ddp",
                market="Australia",
                zone=cells[0],
                channel_name=current_channel or "AU Air DDP",
                source_file=path.name,
                source_sheet=au_air.title,
                transit=cells[5],
                weights=[("11-100KG", 11, 100), ("101-300KG", 101, 300), ("301-500KG", 301, 500), ("501KG+", 501, None)],
                prices=[cells[1], cells[2], cells[3], cells[4]],
            )

    au_sea = wb.worksheets[4]
    current_channel = ""
    for row in au_sea.iter_rows(min_row=1, max_row=30, values_only=True):
        cells = as_cells(row, 10)
        if "澳洲海派" in cells[0]:
            current_channel = cells[0]
        elif cells[0] and (cells[0][0].isdigit() or cells[0] == "其他邮编"):
            add_weight_bands(
                records,
                service_mode="sea_ddp",
                market="Australia",
                zone=cells[0],
                channel_name=current_channel or "AU Sea DDP",
                source_file=path.name,
                source_sheet=au_sea.title,
                transit=cells[9],
                weights=[("11-50KG", 11, 50), ("51-100KG", 51, 100), ("101-300KG", 101, 300), ("301-500KG", 301, 500), ("501-1000KG", 501, 1000), ("1001KG+", 1001, None)],
                prices=[cells[3], cells[4], cells[5], cells[6], cells[7], cells[8]],
            )


def main():
    records = []
    parse_rn(records)
    parse_ythq(records)
    parse_hyjm(records)
    parse_yh_pc(records)
    parse_ym(records)
    parse_yt(records)
    parse_yx(records)

    catalog = {
        "updated_at": "2026-06-26",
        "description": "Generated DDP rate catalog from weekly pricing workbooks.",
        "records": records,
    }
    OUTPUT_PATH.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Records: {len(records)}")


if __name__ == "__main__":
    main()
