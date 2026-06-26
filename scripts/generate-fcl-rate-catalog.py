import json
import re
from pathlib import Path

import openpyxl


OUTPUT_PATH = Path("src/data/fclRateCatalog.generated.json")
SOURCE_DIR = Path(r"C:\Users\YOYO\Documents\WXWork\1688851834237478\Cache\File\2026-06")
SOURCE_SHEET = "海运价格导入模板"

CHINA_PORT_CITY_MAP = {
    "YANTIAN": "Shenzhen",
    "SHEKOU": "Shenzhen",
    "NANSHA": "Guangzhou",
    "HUANGPU": "Guangzhou",
    "DACHAN BAY": "Shenzhen",
    "CHIWAN": "Shenzhen",
    "NINGBO": "Ningbo",
    "SHANGHAI": "Shanghai",
    "QINGDAO": "Qingdao",
    "TIANJINXINGANG": "Tianjin",
    "TIANJIN XINGANG": "Tianjin",
    "TIANJIN": "Tianjin",
    "XIAMEN": "Xiamen",
    "DALIAN": "Dalian",
    "LIANYUNGANG": "Lianyungang",
    "TAICANG": "Taicang",
    "FUZHOU": "Fuzhou",
    "YINGKOU": "Yingkou",
    "WEIHAI": "Weihai",
    "YANTAI": "Yantai",
    "RIZHAO": "Rizhao",
    "ZHANJIANG": "Zhanjiang",
    "ZHONGSHAN": "Zhongshan",
    "ZHUHAI": "Zhuhai",
    "WENZHOU": "Wenzhou",
    "QUANZHOU": "Quanzhou",
    "HONG KONG": "Hong Kong",
}

PORT_NAME_NORMALIZATION = {
    "TIANJINXINGANG": "Tianjin Xingang",
    "TIUANJINXINGANG": "Tianjin Xingang",
    "JEBELALI": "Jebel Ali",
    "KINGABDULLAH": "King Abdullah Port",
    "LOSANGELES": "Los Angeles",
    "LONGBEACH": "Long Beach",
    "NEWYORK": "New York",
}

COUNTRY_NORMALIZATION = {
    "AMERICA": "United States",
    "UNITED STATES": "United States",
    "USA": "United States",
    "CANADA": "Canada",
    "MEXICO": "Mexico",
    "PANAMA": "Panama",
    "COLOMBIA": "Colombia",
    "PERU": "Peru",
    "ARGENTINA": "Argentina",
    "BRAZIL": "Brazil",
    "ITALY": "Italy",
    "CHILE": "Chile",
    "PARAGUAY": "Paraguay",
    "EGYPT": "Egypt",
    "ISRAEL": "Israel",
    "DENMARK": "Denmark",
    "DENMARKTHE KINGDOM OF DENMARK": "Denmark",
    "GERMANY": "Germany",
    "NETHERLANDS": "Netherlands",
    "UNITED KINGDOM": "United Kingdom",
    "UK": "United Kingdom",
    "BELGIUM": "Belgium",
    "FRANCE": "France",
    "SPAIN": "Spain",
    "PORTUGAL": "Portugal",
    "GREECE": "Greece",
    "POLAND": "Poland",
    "TURKIYE": "Turkey",
    "TURKEY": "Turkey",
    "UAE": "United Arab Emirates",
    "UNITED ARAB EMIRATES": "United Arab Emirates",
    "SAUDI ARABIA": "Saudi Arabia",
    "SOUTH AFRICA": "South Africa",
    "NIGERIA": "Nigeria",
    "KENYA": "Kenya",
    "TANZANIA": "Tanzania",
    "MOROCCO": "Morocco",
    "ALGERIA": "Algeria",
    "BAHRAIN": "Bahrain",
    "JORDAN": "Jordan",
    "SINGAPORE": "Singapore",
    "MALAYSIA": "Malaysia",
    "INDIA": "India",
    "JAPAN": "Japan",
    "SOUTH KOREA": "South Korea",
    "BANGLADESH": "Bangladesh",
    "SRI LANKA": "Sri Lanka",
    "PAKISTAN": "Pakistan",
    "THAILAND": "Thailand",
    "BAHAMAS": "Bahamas",
    "CYPRUS": "Cyprus",
    "AUSTRALIA": "Australia",
    "LEBANON": "Lebanon",
    "OMAN": "Oman",
    "QATAR": "Qatar",
    "KUWAIT": "Kuwait",
    "GHANA": "Ghana",
    "ANGOLA": "Angola",
    "DJIBOUTI": "Djibouti",
    "MOZAMBIQUE": "Mozambique",
    "IVORY COAST": "Ivory Coast",
    "COTE D IVOIRE": "Ivory Coast",
}

ADDITIONAL_CHINA_ORIGINS = [
    ("Dalian", "Dalian", "China"),
    ("Fuzhou", "Fuzhou", "China"),
    ("Guangzhou", "Guangzhou", "China"),
    ("Hong Kong", "Hong Kong", "China"),
    ("Huangpu", "Guangzhou", "China"),
    ("Lianyungang", "Lianyungang", "China"),
    ("Nansha", "Guangzhou", "China"),
    ("Ningbo", "Ningbo", "China"),
    ("Qingdao", "Qingdao", "China"),
    ("Rizhao", "Rizhao", "China"),
    ("Shanghai", "Shanghai", "China"),
    ("Shekou", "Shenzhen", "China"),
    ("Taicang", "Taicang", "China"),
    ("Tianjin Xingang", "Tianjin", "China"),
    ("Wenzhou", "Wenzhou", "China"),
    ("Weihai", "Weihai", "China"),
    ("Xiamen", "Xiamen", "China"),
    ("Yantian", "Shenzhen", "China"),
    ("Yantai", "Yantai", "China"),
    ("Yingkou", "Yingkou", "China"),
    ("Zhuhai", "Zhuhai", "China"),
    ("Zhanjiang", "Zhanjiang", "China"),
    ("Quanzhou", "Quanzhou", "China"),
    ("Dachan Bay", "Shenzhen", "China"),
    ("Chiwan", "Shenzhen", "China"),
]

ADDITIONAL_DESTINATIONS = [
    ("Abidjan", "Abidjan", "Ivory Coast"),
    ("Abu Dhabi", "Abu Dhabi", "United Arab Emirates"),
    ("Aqaba", "Aqaba", "Jordan"),
    ("Alexandria", "Alexandria", "Egypt"),
    ("Algeciras", "Algeciras", "Spain"),
    ("Algiers", "Algiers", "Algeria"),
    ("Angoda", "Colombo", "Sri Lanka"),
    ("Antwerp", "Antwerp", "Belgium"),
    ("Ashdod", "Ashdod", "Israel"),
    ("Atlanta", "Atlanta", "United States"),
    ("Balboa", "Balboa", "Panama"),
    ("Baltimore", "Baltimore", "United States"),
    ("Barcelona", "Barcelona", "Spain"),
    ("Beirut", "Beirut", "Lebanon"),
    ("Bilbao", "Bilbao", "Spain"),
    ("Boston", "Boston", "United States"),
    ("Bremerhaven", "Bremerhaven", "Germany"),
    ("Buenos Aires", "Buenos Aires", "Argentina"),
    ("Buenaventura", "Buenaventura", "Colombia"),
    ("Busan", "Busan", "South Korea"),
    ("Callao", "Callao", "Peru"),
    ("Calgary", "Calgary", "Canada"),
    ("Cape Town", "Cape Town", "South Africa"),
    ("Casablanca", "Casablanca", "Morocco"),
    ("Charleston", "Charleston", "United States"),
    ("Chicago", "Chicago", "United States"),
    ("Chittagong", "Chattogram", "Bangladesh"),
    ("Cincinnati", "Cincinnati", "United States"),
    ("Cleveland", "Cleveland", "United States"),
    ("Colombo", "Colombo", "Sri Lanka"),
    ("Columbus", "Columbus", "United States"),
    ("Copenhagen", "Copenhagen", "Denmark"),
    ("Dallas", "Dallas", "United States"),
    ("Dammam", "Dammam", "Saudi Arabia"),
    ("Dar Es Salaam", "Dar es Salaam", "Tanzania"),
    ("Denver", "Denver", "United States"),
    ("Detroit", "Detroit", "United States"),
    ("Djibouti", "Djibouti", "Djibouti"),
    ("Doha", "Doha", "Qatar"),
    ("Dubai", "Dubai", "United Arab Emirates"),
    ("Durban", "Durban", "South Africa"),
    ("Edmonton", "Edmonton", "Canada"),
    ("Felixstowe", "Felixstowe", "United Kingdom"),
    ("Freeport", "Freeport", "Bahamas"),
    ("Gdansk", "Gdansk", "Poland"),
    ("Genoa", "Genoa", "Italy"),
    ("Halifax", "Halifax", "Canada"),
    ("Hamburg", "Hamburg", "Germany"),
    ("Houston", "Houston", "United States"),
    ("Indianapolis", "Indianapolis", "United States"),
    ("Istanbul", "Istanbul", "Turkey"),
    ("Jebel Ali", "Dubai", "United Arab Emirates"),
    ("Jeddah", "Jeddah", "Saudi Arabia"),
    ("Johannesburg", "Johannesburg", "South Africa"),
    ("Karachi", "Karachi", "Pakistan"),
    ("Kuwait", "Kuwait City", "Kuwait"),
    ("Laem Chabang", "Laem Chabang", "Thailand"),
    ("Lagos", "Lagos", "Nigeria"),
    ("Le Havre", "Le Havre", "France"),
    ("Limassol", "Limassol", "Cyprus"),
    ("Lisbon", "Lisbon", "Portugal"),
    ("Liverpool", "Liverpool", "United Kingdom"),
    ("Long Beach", "Long Beach", "United States"),
    ("Los Angeles", "Los Angeles", "United States"),
    ("Louisville", "Louisville", "United States"),
    ("Luanda", "Luanda", "Angola"),
    ("Manzanillo", "Manzanillo", "Mexico"),
    ("Melbourne", "Melbourne", "Australia"),
    ("Memphis", "Memphis", "United States"),
    ("Mersin", "Mersin", "Turkey"),
    ("Miami", "Miami", "United States"),
    ("Milwaukee", "Milwaukee", "United States"),
    ("Mombasa", "Mombasa", "Kenya"),
    ("Montreal", "Montreal", "Canada"),
    ("Mundra", "Mundra", "India"),
    ("Muscat", "Muscat", "Oman"),
    ("Nairobi", "Nairobi", "Kenya"),
    ("Naples", "Naples", "Italy"),
    ("Nashville", "Nashville", "United States"),
    ("New York", "New York", "United States"),
    ("Nhava Sheva", "Mumbai", "India"),
    ("Norfolk", "Norfolk", "United States"),
    ("Oakland", "Oakland", "United States"),
    ("Omaha", "Omaha", "United States"),
    ("Orlando", "Orlando", "United States"),
    ("Osaka", "Osaka", "Japan"),
    ("Ottawa", "Ottawa", "Canada"),
    ("Philadelphia", "Philadelphia", "United States"),
    ("Phoenix", "Phoenix", "United States"),
    ("Piraeus", "Piraeus", "Greece"),
    ("Port Klang", "Port Klang", "Malaysia"),
    ("Port Sudan", "Port Sudan", "Sudan"),
    ("Portland", "Portland", "United States"),
    ("Prince Rupert", "Prince Rupert", "Canada"),
    ("Riyadh Dry Port", "Riyadh", "Saudi Arabia"),
    ("Rotterdam", "Rotterdam", "Netherlands"),
    ("Salt Lake City", "Salt Lake City", "United States"),
    ("San Antonio", "San Antonio", "United States"),
    ("San Diego", "San Diego", "United States"),
    ("San Francisco", "San Francisco", "United States"),
    ("Savannah", "Savannah", "United States"),
    ("Seattle", "Seattle", "United States"),
    ("Singapore", "Singapore", "Singapore"),
    ("Southampton", "Southampton", "United Kingdom"),
    ("St. Louis", "St. Louis", "United States"),
    ("Sydney", "Sydney", "Australia"),
    ("Tacoma", "Tacoma", "United States"),
    ("Tema", "Tema", "Ghana"),
    ("Tokyo", "Tokyo", "Japan"),
    ("Toronto", "Toronto", "Canada"),
    ("Valencia", "Valencia", "Spain"),
    ("Vancouver", "Vancouver", "Canada"),
    ("Vitoria", "Vitoria", "Brazil"),
    ("Winnipeg", "Winnipeg", "Canada"),
]


def split_cn_label(value):
    text = str(value or "").strip()
    if not text:
        return "", ""
    text = text.replace("，", "|").replace("锛", "|").replace("閿", "|").replace("闁", "|")
    english = text.split("|", 1)[0].strip()
    english = re.sub(r"[\u4e00-\u9fff]+", "", english)
    english = re.sub(r"\s+", " ", english).strip(" ,;/|-")
    return english, text


def title_port(value):
    cleaned = str(value or "").replace(" PORT", "").replace(" Port", "").strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    titled = cleaned.title()
    return PORT_NAME_NORMALIZATION.get(titled.replace(" ", "").upper(), titled)


def canonical_key(port, city, country):
    return re.sub(r"[^A-Z0-9]", "", f"{port}|{city}|{country}".upper())


def normalize_country(value):
    key = re.sub(r"[^A-Z ]", "", str(value or "").strip().upper())
    return COUNTRY_NORMALIZATION.get(key, str(value or "").strip().title())


def build_label(port, city, country):
    return f"{port} - {city} - {country}"


def main():
    rates = []
    china_port_map = {}
    international_port_map = {}
    source_files = sorted(SOURCE_DIR.glob("*.xlsx"))

    for path in source_files:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if SOURCE_SHEET not in workbook.sheetnames:
            continue
        sheet = workbook[SOURCE_SHEET]

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            origin_country_raw, origin_port_raw, destination_country_raw, destination_port_raw, carrier, price_20gp, price_40gp, price_40hq = row[:8]
            if not origin_port_raw or not destination_port_raw:
                continue

            origin_country_en, _ = split_cn_label(origin_country_raw or "CHINA")
            origin_port_en, _ = split_cn_label(origin_port_raw)
            destination_country_en, _ = split_cn_label(destination_country_raw)
            destination_port_en, _ = split_cn_label(destination_port_raw)

            origin_port_display = title_port(origin_port_en)
            destination_port_display = title_port(destination_port_en)
            origin_city = CHINA_PORT_CITY_MAP.get(origin_port_display.upper(), origin_port_display)
            destination_country = normalize_country(destination_country_en)
            destination_city = destination_port_display

            if not origin_port_display or not destination_port_display:
                continue

            origin_label = build_label(origin_port_display, origin_city, "China")
            destination_label = build_label(destination_port_display, destination_city, destination_country)

            china_port_map[canonical_key(origin_port_display, origin_city, "China")] = {
                "label": origin_label,
                "port": origin_port_display,
                "city": origin_city,
                "country": "China",
            }
            international_port_map[canonical_key(destination_port_display, destination_city, destination_country)] = {
                "label": destination_label,
                "port": destination_port_display,
                "city": destination_city,
                "country": destination_country,
            }

            for container_type, base_price in {
                "20GP": price_20gp,
                "40GP": price_40gp,
                "40HQ": price_40hq,
            }.items():
                if base_price in (None, ""):
                    continue
                rates.append(
                    {
                        "rate_id": f"fcl-{len(rates) + 1}",
                        "service_type": "fcl",
                        "quote_page_group": "ocean_fcl",
                        "origin_label": origin_label,
                        "destination_label": destination_label,
                        "origin_port": origin_port_display,
                        "destination_port": destination_port_display,
                        "origin_country": origin_country_en.title(),
                        "destination_country": destination_country,
                        "carrier": str(carrier or "").strip() or "General Carrier",
                        "container_type": container_type,
                        "currency": "USD",
                        "base_price": float(base_price),
                        "source_file": path.name,
                        "source_sheet": SOURCE_SHEET,
                        "source_row": row_index,
                    }
                )

    for port, city, country in ADDITIONAL_CHINA_ORIGINS:
        label = build_label(port, city, country)
        china_port_map[canonical_key(port, city, country)] = {
            "label": label,
            "port": port,
            "city": city,
            "country": country,
        }

    for port, city, country in ADDITIONAL_DESTINATIONS:
        label = build_label(port, city, country)
        international_port_map[canonical_key(port, city, country)] = {
            "label": label,
            "port": port,
            "city": city,
            "country": country,
        }

    catalog = {
        "updated_at": "2026-06-26",
        "description": "Generated FCL rate catalog and port option source for Ocean FCL quote search.",
        "china_origin_ports": sorted(china_port_map.values(), key=lambda item: item["label"]),
        "international_destination_ports": sorted(international_port_map.values(), key=lambda item: item["label"]),
        "rates": rates,
    }

    OUTPUT_PATH.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Origin ports: {len(catalog['china_origin_ports'])}")
    print(f"Destination ports: {len(catalog['international_destination_ports'])}")
    print(f"Rates: {len(catalog['rates'])}")


if __name__ == "__main__":
    main()
